from datetime import datetime
from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.services.excel_export import (
    SnapshotExportContext,
    build_filename,
    build_procurement_workbook,
    build_restock_workbook,
)


@pytest.fixture
def procurement_context() -> SnapshotExportContext:
    return SnapshotExportContext(
        suggestion_id=42,
        snapshot_type="procurement",
        version=1,
        exported_at=datetime(2026, 4, 17, 14, 30, 52),
        exported_by_name="alice",
        note="send to supplier A",
        global_config={
            "target_days": 30,
            "buffer_days": 7,
            "lead_time_days": 14,
            "safety_stock_days": 15,
            "restock_regions": ["US", "GB"],
            "eu_countries": ["DE", "FR"],
            "demand_date": "2026-04-30",
        },
        items=[
            {
                "commodity_sku": "SKU-A",
                "commodity_name": "Item A",
                "main_image_url": "https://img/a.jpg",
                "total_qty": 150,
                "purchase_qty": 80,
                "urgent": True,
                "country_breakdown": {"US": 100, "GB": 50},
                "warehouse_breakdown": {"US": {"WH-1": 60, "WH-2": 40}, "GB": {"WH-5": 50}},
                "restock_dates": {"US": "2026-04-23", "GB": "2026-05-13"},
                "velocity_snapshot": {"US": 1.5, "GB": 0.8},
                "sale_days_snapshot": {"US": 20, "GB": 40},
                "local_stock": {"available": 12, "reserved": 3},
            }
        ],
    )


@pytest.fixture
def restock_context() -> SnapshotExportContext:
    return SnapshotExportContext(
        suggestion_id=42,
        snapshot_type="restock",
        version=2,
        exported_at=datetime(2026, 4, 18, 9, 0, 0),
        exported_by_name="alice",
        note=None,
        global_config={
            "target_days": 30,
            "buffer_days": 7,
            "lead_time_days": 14,
            "safety_stock_days": 15,
            "restock_regions": ["US", "GB"],
            "eu_countries": ["DE", "FR"],
            "demand_date": "2026-04-30",
        },
        items=[
            {
                "commodity_sku": "SKU-A",
                "commodity_name": "Item A",
                "main_image_url": "https://img/a.jpg",
                "total_qty": 150,
                "urgent": True,
                "country_breakdown": {"US": 100, "GB": 50},
                "warehouse_breakdown": {"US": {"WH-1": 60, "WH-2": 40}, "GB": {"WH-5": 50}},
                "restock_dates": {"US": "2026-04-24", "GB": "2026-05-14"},
                "velocity_snapshot": {"US": 1.5, "GB": 0.8},
                "sale_days_snapshot": {"US": 20, "GB": 40},
            }
        ],
    )


def test_procurement_workbook_has_two_sheets(procurement_context):
    wb = build_procurement_workbook(procurement_context)
    assert len(wb.sheetnames) == 2


def test_procurement_sheet_rows(procurement_context):
    wb = build_procurement_workbook(procurement_context)
    ws = wb[wb.sheetnames[1]]
    assert ws.max_row == 2
    headers = [ws.cell(row=1, column=index).value for index in range(1, ws.max_column + 1)]
    assert "采购日期" not in headers
    assert "逾期备注" not in headers
    assert ws.cell(row=2, column=1).value == "SKU-A"
    assert ws.cell(row=2, column=4).value == 80
    assert ws.cell(row=2, column=5).value == 2.3


def test_meta_sheet_contains_demand_deadline(procurement_context):
    wb = build_procurement_workbook(procurement_context)
    meta_ws = wb[wb.sheetnames[0]]
    values = {
        meta_ws.cell(row=row, column=1).value: meta_ws.cell(row=row, column=2).value
        for row in range(1, meta_ws.max_row + 1)
    }
    assert values["需求截止日期"] == "2026-04-30"


def test_restock_workbook_has_four_sheets(restock_context):
    wb = build_restock_workbook(restock_context)
    assert len(wb.sheetnames) == 4


def test_restock_country_and_warehouse_rows(restock_context):
    wb = build_restock_workbook(restock_context)
    country_ws = wb[wb.sheetnames[2]]
    warehouse_ws = wb[wb.sheetnames[3]]
    assert country_ws.max_row == 3
    assert country_ws.max_column == 4
    assert warehouse_ws.max_row == 4
    assert warehouse_ws.max_column == 5
    assert country_ws.cell(row=2, column=4).value == "2026-04-24"
    assert warehouse_ws.cell(row=2, column=5).value == "2026-04-24"


def test_workbook_writes_to_bytes(restock_context):
    wb = build_restock_workbook(restock_context)
    payload = BytesIO()
    wb.save(payload)
    payload.seek(0)
    assert len(payload.getvalue()) > 0
    wb2 = load_workbook(payload)
    assert len(wb2.sheetnames) == 4


def test_filename_format():
    name = build_filename(
        suggestion_id=42,
        version=1,
        exported_at=datetime(2026, 4, 17, 14, 30, 52),
        snapshot_type="procurement",
    )
    assert name == "procurement_42_v1_20260417143052.xlsx"
