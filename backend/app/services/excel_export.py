"""Excel export builders for procurement/restock snapshots."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from typing import Any

from openpyxl import Workbook  # type: ignore[import-untyped]
from openpyxl.styles import Alignment, Font, PatternFill  # type: ignore[import-untyped]
from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]


@dataclass
class SnapshotExportContext:
    suggestion_id: int
    snapshot_type: str
    version: int
    exported_at: datetime
    exported_by_name: str | None
    note: str | None
    global_config: dict[str, Any]
    items: list[dict[str, Any]]


HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")


def _apply_header(ws: Any, row: int, headers: list[str]) -> None:
    for col_idx, text in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=text)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN


def _autosize(ws: Any) -> None:
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            value = row[0].value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
        ws.column_dimensions[letter].width = min(max_len + 2, 50)


def build_filename(
    suggestion_id: int,
    version: int,
    exported_at: datetime,
    snapshot_type: str,
) -> str:
    ts = exported_at.strftime("%Y%m%d%H%M%S")
    return f"{snapshot_type}_{suggestion_id}_v{version}_{ts}.xlsx"


def _build_meta_sheet(wb: Workbook, ctx: SnapshotExportContext) -> None:
    ws = wb.create_sheet("主数据")
    rows = [
        ("建议单 ID", ctx.suggestion_id),
        ("快照类型", ctx.snapshot_type),
        ("版本", f"v{ctx.version}"),
        ("导出时间", ctx.exported_at.strftime("%Y-%m-%d %H:%M:%S")),
        ("导出人", ctx.exported_by_name or "系统"),
        ("备注", ctx.note or ""),
        ("buffer_days", ctx.global_config.get("buffer_days", "")),
        ("target_days", ctx.global_config.get("target_days", "")),
        ("lead_time_days", ctx.global_config.get("lead_time_days", "")),
        ("safety_stock_days", ctx.global_config.get("safety_stock_days", "")),
        ("restock_regions", ", ".join(ctx.global_config.get("restock_regions") or [])),
        ("eu_countries", ", ".join(ctx.global_config.get("eu_countries") or [])),
    ]
    for key, value in rows:
        ws.append([key, value])
    _autosize(ws)


def _purchase_date_note(purchase_date: date | None, exported_at: datetime) -> str:
    if purchase_date is None:
        return ""
    delta = (purchase_date - exported_at.date()).days
    if delta < 0:
        return f"逾期 {abs(delta)} 天"
    if delta == 0:
        return "今日到期"
    return ""


def build_procurement_workbook(ctx: SnapshotExportContext) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    _build_meta_sheet(wb, ctx)

    ws = wb.create_sheet("采购明细")
    _apply_header(
        ws,
        1,
        [
            "SKU",
            "商品名",
            "图片 URL",
            "采购量",
            "采购日期",
            "逾期备注",
            "各国动销合计",
            "本地库存可用+占用",
            "安全库存天数",
        ],
    )
    for item in ctx.items:
        velocity_sum = sum((item.get("velocity_snapshot") or {}).values())
        local_stock = item.get("local_stock") or {}
        purchase_date = item.get("purchase_date")
        ws.append(
            [
                item["commodity_sku"],
                item.get("commodity_name") or "",
                item.get("main_image_url") or "",
                item.get("purchase_qty") or 0,
                purchase_date.isoformat() if purchase_date else "",
                _purchase_date_note(purchase_date, ctx.exported_at),
                velocity_sum,
                int(local_stock.get("available", 0)) + int(local_stock.get("reserved", 0)),
                ctx.global_config.get("safety_stock_days", ""),
            ]
        )
    _autosize(ws)
    return wb


def build_restock_workbook(ctx: SnapshotExportContext) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    _build_meta_sheet(wb, ctx)

    sku_ws = wb.create_sheet("SKU汇总")
    _apply_header(sku_ws, 1, ["SKU", "商品名", "补货总量", "紧急"])
    for item in ctx.items:
        sku_ws.append(
            [
                item["commodity_sku"],
                item.get("commodity_name") or "",
                item["total_qty"],
                "是" if item.get("urgent") else "",
            ]
        )
    _autosize(sku_ws)

    country_ws = wb.create_sheet("SKU×国家")
    _apply_header(country_ws, 1, ["SKU", "国家", "补货量"])
    for item in ctx.items:
        for country, qty in (item.get("country_breakdown") or {}).items():
            country_ws.append([item["commodity_sku"], country, qty])
    _autosize(country_ws)

    warehouse_ws = wb.create_sheet("SKU×国家×仓库")
    _apply_header(warehouse_ws, 1, ["SKU", "国家", "仓库", "补货量"])
    for item in ctx.items:
        for country, warehouse_map in (item.get("warehouse_breakdown") or {}).items():
            for warehouse_id, qty in warehouse_map.items():
                warehouse_ws.append([item["commodity_sku"], country, warehouse_id, qty])
    _autosize(warehouse_ws)
    return wb
