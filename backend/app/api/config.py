"""配置管理 API(covers global / sku / warehouse / zipcode / shop)。"""

import csv
from io import BytesIO, StringIO
from typing import Any

from fastapi import APIRouter, Depends, Path, Query, Request
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook  # type: ignore[import-untyped]
from sqlalchemy import delete, distinct, func, or_, select, update
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.api.deps import (
    UserContext,
    db_session,
    db_session_readonly,
    get_current_user,
    require_permission,
)
from app.core.countries import (
    BUILTIN_COUNTRY_NAMES,
    BUILTIN_COUNTRY_ORDER,
    NON_EU_MEMBER_CODES,
    country_label,
    normalize_observed_country_code,
)
from app.core.country_mapping import backfill_eu_country_mapping
from app.core.exceptions import ConflictError, NotFound, UnprocessableError, ValidationFailed
from app.core.permissions import (
    CONFIG_EDIT,
    CONFIG_VIEW,
    DATA_BASE_EDIT,
    DATA_BASE_VIEW,
    RESTOCK_NEW_CYCLE,
    SYNC_OPERATE,
)
from app.core.query import escape_like
from app.core.timezone import now_beijing
from app.models.commodity import CommodityMaster
from app.models.dashboard_snapshot import DashboardSnapshot
from app.models.global_config import GlobalConfig
from app.models.in_transit import InTransitRecord
from app.models.inventory import InventorySnapshotLatest
from app.models.order import OrderHeader
from app.models.physical_item import PhysicalItemGroup, PhysicalItemSkuAlias
from app.models.product_listing import ProductListing
from app.models.shop import Shop
from app.models.sku import SkuConfig
from app.models.sku_mapping import SkuMappingComponent, SkuMappingRule
from app.models.suggestion import Suggestion
from app.models.suggestion_snapshot import SuggestionSnapshot
from app.models.sys_user import SysUser
from app.models.warehouse import Warehouse
from app.models.zipcode_rule import ZipcodeRule
from app.schemas.config import (
    CountryOptionOut,
    CountryOptionsOut,
    GenerationToggleOut,
    GenerationTogglePatch,
    GlobalConfigOut,
    GlobalConfigPatch,
    PhysicalItemGroupIn,
    PhysicalItemGroupListOut,
    PhysicalItemGroupOut,
    PhysicalItemGroupPatch,
    ShopOut,
    ShopPatch,
    SkuConfigListOut,
    SkuConfigOut,
    SkuConfigPatch,
    SkuMappingComponentIn,
    SkuMappingImportOut,
    SkuMappingRuleIn,
    SkuMappingRuleListOut,
    SkuMappingRuleOut,
    SkuMappingRulePatch,
    WarehouseCountryPatch,
    WarehouseOut,
    ZipcodeRuleIn,
    ZipcodeRuleOut,
)
from app.services.physical_item import load_physical_sku_resolver
from app.tasks.queue import enqueue_task
from app.tasks.scheduler import reload_scheduler

router = APIRouter(prefix="/api/config", tags=["config"])

# 改动以下任一字段即把 dashboard_snapshot.stale 置 TRUE，下次 dashboard API
# 自动 enqueue 刷新。
GLOBAL_CONFIG_SENSITIVE_FIELDS = frozenset(
    {
        "restock_regions",
        "eu_countries",
        "target_days",
        "lead_time_days",
        "buffer_days",
        "safety_stock_days",
    }
)

# 与 dashboard 展示无关、变更无需 stale 的字段。新增 GlobalConfig 字段
# 必须在 SENSITIVE 或 NEUTRAL 任一集合中声明，否则 tripwire 测试会红。
GLOBAL_CONFIG_NEUTRAL_FIELDS = frozenset(
    {
        "id",
        "sync_interval_minutes",  # 调度触发频率，不影响 dashboard 数据
        "scheduler_enabled",
        "shop_sync_mode",
        "login_password_hash",
        "suggestion_generation_enabled",
        "generation_toggle_updated_by",
        "generation_toggle_updated_at",
        "created_at",
        "updated_at",
    }
)


def _warehouse_total_stock_subquery() -> Any:
    return (
        select(
            InventorySnapshotLatest.warehouse_id.label("warehouse_id"),
            func.coalesce(
                func.sum(InventorySnapshotLatest.available + InventorySnapshotLatest.reserved),
                0,
            ).label("total_stock"),
        )
        .group_by(InventorySnapshotLatest.warehouse_id)
        .subquery()
    )


def _warehouse_list_stmt() -> Any:
    stock_subquery = _warehouse_total_stock_subquery()
    return (
        select(
            Warehouse,
            func.coalesce(stock_subquery.c.total_stock, 0).label("total_stock"),
        )
        .outerjoin(stock_subquery, stock_subquery.c.warehouse_id == Warehouse.id)
        .order_by(Warehouse.country, Warehouse.id)
    )


def _warehouse_out_from_row(row: Any) -> WarehouseOut:
    warehouse, total_stock = row
    return WarehouseOut.model_validate(
        {
            "id": warehouse.id,
            "name": warehouse.name,
            "type": warehouse.type,
            "country": warehouse.country,
            "replenish_site_raw": warehouse.replenish_site_raw,
            "total_stock": int(total_stock or 0),
        }
    )


async def init_sku_configs_from_listings(db: AsyncSession) -> int:
    sku_codes = sorted(
        (await db.execute(select(CommodityMaster.sku).order_by(CommodityMaster.sku)))
        .scalars()
        .all()
    )
    if not sku_codes:
        sku_codes = sorted(
            {
                sku
                for sku in (
                    await db.execute(
                        select(ProductListing.commodity_sku)
                        .where(ProductListing.commodity_sku.is_not(None))
                        .order_by(ProductListing.commodity_sku)
                    )
                )
                .scalars()
                .all()
                if sku is not None
            }
        )
    if not sku_codes:
        return 0

    existing_codes = set(
        (
            await db.execute(
                select(SkuConfig.commodity_sku).where(SkuConfig.commodity_sku.in_(sku_codes))
            )
        )
        .scalars()
        .all()
    )
    missing_codes = [code for code in sku_codes if code not in existing_codes]
    if not missing_codes:
        return 0

    await db.execute(
        pg_insert(SkuConfig).values(
            [{"commodity_sku": code, "enabled": True} for code in missing_codes]
        )
    )
    await db.commit()
    return len(missing_codes)


# ============================================================
# Global Config
# ============================================================
@router.get("/global", response_model=GlobalConfigOut)
async def get_global(
    db: AsyncSession = Depends(db_session_readonly),
    _: None = Depends(require_permission(CONFIG_VIEW)),
) -> GlobalConfigOut:
    row = (await db.execute(select(GlobalConfig).where(GlobalConfig.id == 1))).scalar_one()
    return GlobalConfigOut.model_validate(row)


async def _observed_country_codes(db: AsyncSession) -> set[str]:
    columns = (
        OrderHeader.country_code,
        OrderHeader.original_country_code,
        Warehouse.country,
        InventorySnapshotLatest.country,
        InventorySnapshotLatest.original_country,
        InTransitRecord.target_country,
        InTransitRecord.original_target_country,
    )
    observed: set[str] = set()
    for column in columns:
        rows = (
            (
                await db.execute(
                    select(distinct(column)).where(column.is_not(None)).where(column != "")
                )
            )
            .scalars()
            .all()
        )
        for raw in rows:
            code = normalize_observed_country_code(raw)
            if code is not None:
                observed.add(code)
    return observed


@router.get("/country-options", response_model=CountryOptionsOut)
async def get_country_options(
    db: AsyncSession = Depends(db_session_readonly),
    _: None = Depends(require_permission(CONFIG_VIEW)),
) -> CountryOptionsOut:
    observed = await _observed_country_codes(db)
    all_codes = set(BUILTIN_COUNTRY_NAMES) | observed
    builtin_set = set(BUILTIN_COUNTRY_ORDER)
    ordered_codes = [
        *[code for code in BUILTIN_COUNTRY_ORDER if code in all_codes],
        *sorted(all_codes - builtin_set),
    ]
    unknown_codes = sorted(
        code
        for code in observed
        if code not in BUILTIN_COUNTRY_NAMES and code not in NON_EU_MEMBER_CODES
    )
    return CountryOptionsOut(
        items=[
            CountryOptionOut(
                code=code,
                label=country_label(code),
                builtin=code in BUILTIN_COUNTRY_NAMES,
                observed=code in observed,
                can_be_eu_member=code not in NON_EU_MEMBER_CODES,
            )
            for code in ordered_codes
        ],
        unknown_country_codes=unknown_codes,
    )


@router.patch("/global", response_model=GlobalConfigOut)
async def patch_global(
    patch: GlobalConfigPatch,
    db: AsyncSession = Depends(db_session),
    _: None = Depends(require_permission(CONFIG_EDIT)),
) -> GlobalConfigOut:
    row = (await db.execute(select(GlobalConfig).where(GlobalConfig.id == 1))).scalar_one()
    updates = patch.model_dump(exclude_none=True)
    if updates:
        target_days = updates.get("target_days", row.target_days)
        lead_time_days = updates.get("lead_time_days", row.lead_time_days)
        if target_days < lead_time_days:
            raise ValidationFailed("目标库存天数不能小于采购提前期")
        # 值感知前先 snapshot 旧值——一旦 execute(update(...)) 触发 ORM
        # auto-expire，后续 getattr(row, field) 会懒加载拿到新值，无法判断变更。
        sensitive_updates = GLOBAL_CONFIG_SENSITIVE_FIELDS & updates.keys()
        sensitive_old = {f: getattr(row, f, None) for f in sensitive_updates}
        await db.execute(update(GlobalConfig).where(GlobalConfig.id == 1).values(**updates))
        if (
            "eu_countries" in updates
            and sensitive_old.get("eu_countries") != updates["eu_countries"]
        ):
            await backfill_eu_country_mapping(db, updates["eu_countries"])
        if sensitive_updates and any(sensitive_old[f] != updates[f] for f in sensitive_updates):
            await db.execute(
                update(DashboardSnapshot).where(DashboardSnapshot.id == 1).values(stale=True)
            )
        await db.commit()
        if {"sync_interval_minutes", "scheduler_enabled"} & updates.keys():
            await reload_scheduler()
        row = (await db.execute(select(GlobalConfig).where(GlobalConfig.id == 1))).scalar_one()
    return GlobalConfigOut.model_validate(row)


# ============================================================
# Generation Toggle（补货建议生成总开关）
# ============================================================
async def _compute_can_enable(db: AsyncSession) -> tuple[bool, str | None]:
    draft = (
        await db.execute(
            select(Suggestion)
            .where(Suggestion.status == "draft")
            .order_by(Suggestion.created_at.desc())
            .limit(1)
        )
    ).scalar_one_or_none()
    if draft is None:
        return True, None

    if draft.procurement_item_count > 0:
        procurement_count = (
            await db.execute(
                select(func.count())
                .select_from(SuggestionSnapshot)
                .where(
                    SuggestionSnapshot.suggestion_id == draft.id,
                    SuggestionSnapshot.snapshot_type == "procurement",
                )
            )
        ).scalar_one()
        if int(procurement_count or 0) == 0:
            return False, "采购建议尚未导出任何快照"

    if draft.restock_item_count > 0:
        restock_count = (
            await db.execute(
                select(func.count())
                .select_from(SuggestionSnapshot)
                .where(
                    SuggestionSnapshot.suggestion_id == draft.id,
                    SuggestionSnapshot.snapshot_type == "restock",
                )
            )
        ).scalar_one()
        if int(restock_count or 0) == 0:
            return False, "补货建议尚未导出任何快照"

    return True, None


async def _load_generation_toggle(db: AsyncSession) -> GenerationToggleOut:
    row = (
        await db.execute(
            select(
                GlobalConfig.suggestion_generation_enabled,
                GlobalConfig.generation_toggle_updated_by,
                GlobalConfig.generation_toggle_updated_at,
                SysUser.display_name,
            )
            .outerjoin(SysUser, SysUser.id == GlobalConfig.generation_toggle_updated_by)
            .where(GlobalConfig.id == 1)
        )
    ).one()
    enabled, updated_by, updated_at, display_name = row
    can_enable, can_enable_reason = await _compute_can_enable(db)
    return GenerationToggleOut(
        enabled=bool(enabled),
        updated_by=updated_by,
        updated_by_name=display_name,
        updated_at=updated_at,
        can_enable=can_enable,
        can_enable_reason=can_enable_reason,
    )


@router.get("/generation-toggle", response_model=GenerationToggleOut)
async def get_generation_toggle(
    db: AsyncSession = Depends(db_session_readonly),
    _: None = Depends(require_permission(CONFIG_VIEW)),
) -> GenerationToggleOut:
    return await _load_generation_toggle(db)


@router.patch("/generation-toggle", response_model=GenerationToggleOut)
async def patch_generation_toggle(
    patch: GenerationTogglePatch,
    db: AsyncSession = Depends(db_session),
    user: UserContext = Depends(get_current_user),
    _: None = Depends(require_permission(RESTOCK_NEW_CYCLE)),
) -> GenerationToggleOut:
    """翻开关。若打开（enabled=True）: 先归档所有 status='draft' 的 suggestion。"""
    now = now_beijing()
    config = (
        await db.execute(select(GlobalConfig).where(GlobalConfig.id == 1).with_for_update())
    ).scalar_one()
    if patch.enabled:
        can_enable, reason = await _compute_can_enable(db)
        if not can_enable:
            raise UnprocessableError(reason or "无法开启：前置条件不满足")
        await db.execute(
            update(Suggestion)
            .where(Suggestion.status == "draft")
            .values(
                status="archived",
                archived_at=now,
                archived_by=user.id,
                archived_trigger="admin_toggle",
            )
        )
    config.suggestion_generation_enabled = patch.enabled
    config.generation_toggle_updated_by = user.id
    config.generation_toggle_updated_at = now
    await db.commit()
    return await _load_generation_toggle(db)


# ============================================================
# SKU Config
# ============================================================
@router.get("/sku", response_model=SkuConfigListOut)
async def list_sku_configs(
    enabled: bool | None = Query(default=None),
    keyword: str | None = Query(default=None),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=50, ge=1, le=200),
    db: AsyncSession = Depends(db_session_readonly),
    _: None = Depends(require_permission(DATA_BASE_VIEW)),
) -> SkuConfigListOut:
    base = select(SkuConfig).order_by(SkuConfig.commodity_sku)
    if enabled is not None:
        base = base.where(SkuConfig.enabled.is_(enabled))
    if keyword:
        base = base.where(SkuConfig.commodity_sku.ilike(f"%{escape_like(keyword)}%", escape="\\"))

    total = (await db.execute(select(func.count()).select_from(base.subquery()))).scalar_one()
    rows = (await db.execute(base.offset((page - 1) * page_size).limit(page_size))).scalars().all()

    # 一次性 JOIN product_listing 获取 commodity_name + main_image
    sku_codes = [r.commodity_sku for r in rows]
    name_map: dict[str, tuple[str | None, str | None]] = {}
    if sku_codes:
        commodity_rows = (
            await db.execute(
                select(
                    CommodityMaster.sku,
                    CommodityMaster.name,
                    CommodityMaster.img_url,
                ).where(CommodityMaster.sku.in_(sku_codes))
            )
        ).all()
        for sku, name, img in commodity_rows:
            name_map[sku] = (name, img)
        listing_rows = (
            await db.execute(
                select(
                    ProductListing.commodity_sku,
                    ProductListing.commodity_name,
                    ProductListing.main_image,
                ).where(ProductListing.commodity_sku.in_(sku_codes))
            )
        ).all()
        for sku, name, img in listing_rows:
            if sku not in name_map:
                name_map[sku] = (name, img)

    items = []
    for r in rows:
        name, img = name_map.get(r.commodity_sku, (None, None))
        items.append(
            SkuConfigOut(
                commodity_sku=r.commodity_sku,
                enabled=r.enabled,
                lead_time_days=r.lead_time_days,
                commodity_name=name,
                main_image=img,
            )
        )
    return SkuConfigListOut(items=items, total=int(total or 0))


@router.post("/sku/init")
async def init_sku_configs(
    db: AsyncSession = Depends(db_session),
    _: None = Depends(require_permission(DATA_BASE_EDIT)),
) -> dict[str, int]:
    created = await init_sku_configs_from_listings(db)
    total = (await db.execute(select(func.count()).select_from(SkuConfig))).scalar_one()
    return {"created": created, "total": int(total or 0)}


@router.patch("/sku/{commodity_sku}", response_model=SkuConfigOut)
async def patch_sku_config(
    patch: SkuConfigPatch,
    commodity_sku: str = Path(...),
    db: AsyncSession = Depends(db_session),
    _: None = Depends(require_permission(DATA_BASE_EDIT)),
) -> SkuConfigOut:
    row = (
        await db.execute(select(SkuConfig).where(SkuConfig.commodity_sku == commodity_sku))
    ).scalar_one_or_none()
    if row is None:
        raise NotFound(f"SKU {commodity_sku} 未配置")
    updates = patch.model_dump(exclude_unset=True)
    if updates:
        await db.execute(
            update(SkuConfig).where(SkuConfig.commodity_sku == commodity_sku).values(**updates)
        )
        await db.refresh(row)
    return SkuConfigOut.model_validate({**row.__dict__, "commodity_name": None, "main_image": None})


# ============================================================
# SKU Mapping Rules
# ============================================================
def _mapping_formula(rule: SkuMappingRule) -> str:
    groups: dict[int, list[SkuMappingComponent]] = {}
    for component in rule.components:
        groups.setdefault(component.group_no, []).append(component)
    group_parts = [
        "+".join(
            f"{component.quantity}*{component.inventory_sku}" for component in groups[group_no]
        )
        for group_no in sorted(groups)
    ]
    return f"{rule.commodity_sku}=" + " 或 ".join(group_parts)


def _mapping_rule_out(rule: SkuMappingRule) -> SkuMappingRuleOut:
    return SkuMappingRuleOut(
        id=rule.id,
        commodity_sku=rule.commodity_sku,
        enabled=rule.enabled,
        remark=rule.remark,
        components=rule.components,
        formula_preview=_mapping_formula(rule),
        component_count=len(rule.components),
        created_at=rule.created_at,
        updated_at=rule.updated_at,
    )


async def _validate_mapping_unique(
    db: AsyncSession,
    *,
    commodity_sku: str,
    current_rule_id: int | None = None,
) -> None:
    existing_rule = (
        await db.execute(
            select(SkuMappingRule).where(SkuMappingRule.commodity_sku == commodity_sku)
        )
    ).scalar_one_or_none()
    if existing_rule is not None and existing_rule.id != current_rule_id:
        raise ConflictError(f"商品SKU {commodity_sku} 已存在映射规则")


async def _replace_mapping_components(
    db: AsyncSession,
    rule: SkuMappingRule,
    components: list[SkuMappingComponentIn],
) -> None:
    resolver = await load_physical_sku_resolver(db)
    seen_by_group: dict[int, set[str]] = {}
    for component in components:
        normalized_sku = resolver.resolve_inventory_sku(component.inventory_sku)
        group_seen = seen_by_group.setdefault(component.group_no, set())
        if normalized_sku in group_seen:
            raise ValidationFailed(
                f"同一方案内库存SKU {component.inventory_sku} 与其他组件归属同一库存共享组"
            )
        group_seen.add(normalized_sku)
    await db.execute(delete(SkuMappingComponent).where(SkuMappingComponent.rule_id == rule.id))
    for component in components:
        db.add(
            SkuMappingComponent(
                rule_id=rule.id,
                group_no=component.group_no,
                inventory_sku=component.inventory_sku,
                quantity=component.quantity,
            )
        )


@router.get("/sku-mapping-rules", response_model=SkuMappingRuleListOut)
async def list_sku_mapping_rules(
    enabled: bool | None = Query(default=None),
    keyword: str | None = Query(default=None),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=50, ge=1, le=200),
    db: AsyncSession = Depends(db_session_readonly),
    _: None = Depends(require_permission(CONFIG_VIEW)),
) -> SkuMappingRuleListOut:
    base = select(SkuMappingRule).options(selectinload(SkuMappingRule.components))
    count_stmt = select(func.count(distinct(SkuMappingRule.id))).select_from(SkuMappingRule)
    if enabled is not None:
        base = base.where(SkuMappingRule.enabled.is_(enabled))
        count_stmt = count_stmt.where(SkuMappingRule.enabled.is_(enabled))
    if keyword:
        like = f"%{escape_like(keyword)}%"
        base = base.outerjoin(SkuMappingComponent).where(
            or_(
                SkuMappingRule.commodity_sku.ilike(like, escape="\\"),
                SkuMappingComponent.inventory_sku.ilike(like, escape="\\"),
            )
        )
        count_stmt = count_stmt.outerjoin(SkuMappingComponent).where(
            or_(
                SkuMappingRule.commodity_sku.ilike(like, escape="\\"),
                SkuMappingComponent.inventory_sku.ilike(like, escape="\\"),
            )
        )

    total = (await db.execute(count_stmt)).scalar_one()
    rows = (
        (
            await db.execute(
                base.order_by(SkuMappingRule.commodity_sku)
                .offset((page - 1) * page_size)
                .limit(page_size)
            )
        )
        .scalars()
        .unique()
        .all()
    )
    return SkuMappingRuleListOut(
        items=[_mapping_rule_out(row) for row in rows], total=int(total or 0)
    )


@router.post("/sku-mapping-rules", response_model=SkuMappingRuleOut, status_code=201)
async def create_sku_mapping_rule(
    body: SkuMappingRuleIn,
    db: AsyncSession = Depends(db_session),
    _: None = Depends(require_permission(CONFIG_EDIT)),
) -> SkuMappingRuleOut:
    await _validate_mapping_unique(
        db,
        commodity_sku=body.commodity_sku,
    )
    rule = SkuMappingRule(
        commodity_sku=body.commodity_sku,
        enabled=body.enabled,
        remark=body.remark,
    )
    db.add(rule)
    await db.flush()
    await _replace_mapping_components(db, rule, body.components)
    await db.commit()
    rule = (
        await db.execute(
            select(SkuMappingRule)
            .options(selectinload(SkuMappingRule.components))
            .where(SkuMappingRule.id == rule.id)
        )
    ).scalar_one()
    return _mapping_rule_out(rule)


@router.get("/sku-mapping-rules/export")
async def export_sku_mapping_rules(
    db: AsyncSession = Depends(db_session_readonly),
    _: None = Depends(require_permission(CONFIG_VIEW)),
) -> StreamingResponse:
    rows = (
        (
            await db.execute(
                select(SkuMappingRule)
                .options(selectinload(SkuMappingRule.components))
                .order_by(SkuMappingRule.commodity_sku)
            )
        )
        .scalars()
        .all()
    )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "映射规则"
    sheet.append(["商品SKU", "组合编号", "库存SKU", "组件数量", "启用", "备注"])
    for rule in rows:
        for component in rule.components:
            sheet.append(
                [
                    rule.commodity_sku,
                    component.group_no,
                    component.inventory_sku,
                    component.quantity,
                    "是" if rule.enabled else "否",
                    rule.remark or "",
                ]
            )
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    filename = f"sku_mapping_rules_{now_beijing().strftime('%Y%m%d%H%M%S')}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _parse_enabled(value: Any) -> bool:
    text = str(value if value is not None else "").strip().lower()
    if text in {"", "1", "true", "yes", "y", "是", "启用"}:
        return True
    if text in {"0", "false", "no", "n", "否", "停用"}:
        return False
    raise ValueError("启用列必须为 是/否、true/false 或 1/0")


def _read_mapping_import_rows(filename: str, content: bytes) -> list[dict[str, Any]]:
    filename = filename.lower()
    if filename.endswith(".csv"):
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(StringIO(text))
        return [dict(row) for row in reader]

    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(cell or "").strip() for cell in rows[0]]
    result: list[dict[str, Any]] = []
    for row in rows[1:]:
        result.append(
            {headers[idx]: row[idx] if idx < len(row) else None for idx in range(len(headers))}
        )
    return result


def _normalize_import_rows(raw_rows: list[dict[str, Any]]) -> dict[str, SkuMappingRuleIn]:
    errors: list[dict[str, Any]] = []
    grouped: dict[str, dict[str, Any]] = {}
    seen_inventory: dict[str, dict[str, int]] = {}
    required_headers = {"商品SKU", "库存SKU", "组件数量", "启用", "备注"}

    for idx, raw in enumerate(raw_rows, start=2):
        if not any(str(value or "").strip() for value in raw.values()):
            continue
        missing = required_headers - set(raw)
        if missing:
            errors.append({"row": idx, "message": f"缺少列：{', '.join(sorted(missing))}"})
            continue
        commodity_sku = str(raw.get("商品SKU") or "").strip()
        inventory_sku = str(raw.get("库存SKU") or "").strip()
        group_no_raw = raw.get("组合编号")
        group_no_text = str(group_no_raw or "").strip()
        if not group_no_text:
            group_no = 1
        else:
            try:
                group_no = int(group_no_text)
            except (TypeError, ValueError):
                group_no = 0
        quantity_raw = raw.get("组件数量")
        if quantity_raw is None:
            quantity = 0
        else:
            try:
                quantity = int(quantity_raw)
            except (TypeError, ValueError):
                quantity = 0
        try:
            enabled = _parse_enabled(raw.get("启用"))
        except ValueError as exc:
            errors.append({"row": idx, "message": str(exc)})
            continue
        remark = str(raw.get("备注") or "").strip() or None

        if not commodity_sku:
            errors.append({"row": idx, "message": "商品SKU不能为空"})
        if not inventory_sku:
            errors.append({"row": idx, "message": "库存SKU不能为空"})
        if group_no <= 0:
            errors.append({"row": idx, "message": "组合编号必须为正整数"})
        if quantity <= 0:
            errors.append({"row": idx, "message": "组件数量必须为正整数"})
        if inventory_sku and commodity_sku:
            commodity_seen = seen_inventory.setdefault(commodity_sku, {})
            previous_row = commodity_seen.get(inventory_sku)
            if previous_row is not None:
                errors.append(
                    {"row": idx, "message": f"库存SKU与第 {previous_row} 行重复：{inventory_sku}"}
                )
            commodity_seen[inventory_sku] = idx
        if not commodity_sku or not inventory_sku or group_no <= 0 or quantity <= 0:
            continue

        entry = grouped.setdefault(
            commodity_sku,
            {"enabled": enabled, "remark": remark, "components": []},
        )
        if entry["enabled"] != enabled:
            errors.append({"row": idx, "message": f"同一商品SKU的启用状态不一致：{commodity_sku}"})
        if entry["remark"] != remark:
            errors.append({"row": idx, "message": f"同一商品SKU的备注不一致：{commodity_sku}"})
        entry["components"].append(
            {"group_no": group_no, "inventory_sku": inventory_sku, "quantity": quantity}
        )

    if errors:
        raise ValidationFailed("导入校验失败", detail={"errors": errors})
    return {
        commodity_sku: SkuMappingRuleIn(
            commodity_sku=commodity_sku,
            enabled=entry["enabled"],
            remark=entry["remark"],
            components=entry["components"],
        )
        for commodity_sku, entry in grouped.items()
    }


@router.post("/sku-mapping-rules/import", response_model=SkuMappingImportOut)
async def import_sku_mapping_rules(
    request: Request,
    db: AsyncSession = Depends(db_session),
    _: None = Depends(require_permission(CONFIG_EDIT)),
) -> SkuMappingImportOut:
    filename = request.headers.get("x-filename", "")
    content = await request.body()
    if not content:
        raise ValidationFailed("导入文件没有有效数据")
    raw_rows = _read_mapping_import_rows(filename, content)
    rules = _normalize_import_rows(raw_rows)
    if not rules:
        raise ValidationFailed("导入文件没有有效数据")

    commodity_skus = set(rules)
    component_skus = [
        component.inventory_sku for rule in rules.values() for component in rule.components
    ]
    existing_rows = (
        (
            await db.execute(
                select(SkuMappingRule)
                .options(selectinload(SkuMappingRule.components))
                .where(SkuMappingRule.commodity_sku.in_(commodity_skus))
            )
        )
        .scalars()
        .all()
    )
    existing_map = {row.commodity_sku: row for row in existing_rows}
    created = 0
    updated = 0
    for commodity_sku, body in rules.items():
        rule = existing_map.get(commodity_sku)
        if rule is None:
            rule = SkuMappingRule(
                commodity_sku=commodity_sku,
                enabled=body.enabled,
                remark=body.remark,
            )
            db.add(rule)
            await db.flush()
            created += 1
        else:
            rule.enabled = body.enabled
            rule.remark = body.remark
            updated += 1
        await _replace_mapping_components(db, rule, body.components)
    await db.commit()
    return SkuMappingImportOut(
        created=created,
        updated=updated,
        total_components=len(component_skus),
    )


@router.patch("/sku-mapping-rules/{rule_id}", response_model=SkuMappingRuleOut)
async def patch_sku_mapping_rule(
    patch: SkuMappingRulePatch,
    rule_id: int = Path(..., ge=1),
    db: AsyncSession = Depends(db_session),
    _: None = Depends(require_permission(CONFIG_EDIT)),
) -> SkuMappingRuleOut:
    rule = (
        await db.execute(
            select(SkuMappingRule)
            .options(selectinload(SkuMappingRule.components))
            .where(SkuMappingRule.id == rule_id)
        )
    ).scalar_one_or_none()
    if rule is None:
        raise NotFound(f"映射规则 {rule_id} 不存在")

    new_commodity_sku = (
        patch.commodity_sku if patch.commodity_sku is not None else rule.commodity_sku
    )
    await _validate_mapping_unique(
        db,
        commodity_sku=new_commodity_sku,
        current_rule_id=rule.id,
    )

    if patch.commodity_sku is not None:
        rule.commodity_sku = patch.commodity_sku
    if patch.enabled is not None:
        rule.enabled = patch.enabled
    if "remark" in patch.model_fields_set:
        rule.remark = patch.remark
    if patch.components is not None:
        await _replace_mapping_components(db, rule, patch.components)
    await db.commit()
    rule = (
        await db.execute(
            select(SkuMappingRule)
            .options(selectinload(SkuMappingRule.components))
            .where(SkuMappingRule.id == rule_id)
        )
    ).scalar_one()
    return _mapping_rule_out(rule)


@router.delete("/sku-mapping-rules/{rule_id}", status_code=204)
async def delete_sku_mapping_rule(
    rule_id: int = Path(..., ge=1),
    db: AsyncSession = Depends(db_session),
    _: None = Depends(require_permission(CONFIG_EDIT)),
) -> None:
    result = await db.execute(delete(SkuMappingRule).where(SkuMappingRule.id == rule_id))
    if result.rowcount == 0:  # type: ignore[attr-defined]
        raise NotFound(f"映射规则 {rule_id} 不存在")
    await db.commit()
    return None


# ============================================================
# Physical Item Groups
# ============================================================
def _physical_group_out(group: PhysicalItemGroup) -> PhysicalItemGroupOut:
    return PhysicalItemGroupOut(
        id=group.id,
        name=group.name,
        enabled=group.enabled,
        remark=group.remark,
        members=group.aliases,
        member_count=len(group.aliases),
        created_at=group.created_at,
        updated_at=group.updated_at,
    )


async def _validate_physical_group_unique(
    db: AsyncSession,
    *,
    name: str,
    members: list[str],
    current_group_id: int | None = None,
) -> None:
    name_row = (
        await db.execute(select(PhysicalItemGroup).where(PhysicalItemGroup.name == name))
    ).scalar_one_or_none()
    if name_row is not None and name_row.id != current_group_id:
        raise ConflictError(f"库存SKU共享组名称 {name} 已存在")

    alias_stmt = select(PhysicalItemSkuAlias).where(PhysicalItemSkuAlias.sku.in_(members))
    if current_group_id is not None:
        alias_stmt = alias_stmt.where(PhysicalItemSkuAlias.group_id != current_group_id)
    alias_rows = (await db.execute(alias_stmt)).scalars().all()
    if alias_rows:
        conflict_skus = ", ".join(sorted({row.sku for row in alias_rows}))
        raise ConflictError(f"库存SKU 已归属其他共享组：{conflict_skus}")


async def _replace_physical_members(
    db: AsyncSession,
    group: PhysicalItemGroup,
    members: list[str],
) -> None:
    await db.execute(delete(PhysicalItemSkuAlias).where(PhysicalItemSkuAlias.group_id == group.id))
    for sku in members:
        db.add(PhysicalItemSkuAlias(group_id=group.id, sku=sku))


@router.get("/physical-item-groups", response_model=PhysicalItemGroupListOut)
async def list_physical_item_groups(
    enabled: bool | None = Query(default=None),
    keyword: str | None = Query(default=None),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=50, ge=1, le=500),
    db: AsyncSession = Depends(db_session_readonly),
    _: None = Depends(require_permission(CONFIG_VIEW)),
) -> PhysicalItemGroupListOut:
    base = select(PhysicalItemGroup).options(selectinload(PhysicalItemGroup.aliases))
    count_stmt = select(func.count(distinct(PhysicalItemGroup.id))).select_from(PhysicalItemGroup)
    if enabled is not None:
        base = base.where(PhysicalItemGroup.enabled.is_(enabled))
        count_stmt = count_stmt.where(PhysicalItemGroup.enabled.is_(enabled))
    if keyword:
        like = f"%{escape_like(keyword)}%"
        base = base.outerjoin(PhysicalItemSkuAlias).where(
            or_(
                PhysicalItemGroup.name.ilike(like, escape="\\"),
                PhysicalItemSkuAlias.sku.ilike(like, escape="\\"),
            )
        )
        count_stmt = count_stmt.outerjoin(PhysicalItemSkuAlias).where(
            or_(
                PhysicalItemGroup.name.ilike(like, escape="\\"),
                PhysicalItemSkuAlias.sku.ilike(like, escape="\\"),
            )
        )

    total = (await db.execute(count_stmt)).scalar_one()
    rows = (
        (
            await db.execute(
                base.order_by(PhysicalItemGroup.name)
                .offset((page - 1) * page_size)
                .limit(page_size)
            )
        )
        .scalars()
        .unique()
        .all()
    )
    return PhysicalItemGroupListOut(
        items=[_physical_group_out(row) for row in rows],
        total=int(total or 0),
    )


@router.post("/physical-item-groups", response_model=PhysicalItemGroupOut, status_code=201)
async def create_physical_item_group(
    body: PhysicalItemGroupIn,
    db: AsyncSession = Depends(db_session),
    _: None = Depends(require_permission(CONFIG_EDIT)),
) -> PhysicalItemGroupOut:
    await _validate_physical_group_unique(db, name=body.name, members=body.members)
    group = PhysicalItemGroup(
        name=body.name,
        enabled=body.enabled,
        remark=body.remark,
    )
    db.add(group)
    await db.flush()
    await _replace_physical_members(db, group, body.members)
    await db.commit()
    group = (
        await db.execute(
            select(PhysicalItemGroup)
            .options(selectinload(PhysicalItemGroup.aliases))
            .where(PhysicalItemGroup.id == group.id)
        )
    ).scalar_one()
    return _physical_group_out(group)


@router.patch("/physical-item-groups/{group_id}", response_model=PhysicalItemGroupOut)
async def patch_physical_item_group(
    patch: PhysicalItemGroupPatch,
    group_id: int = Path(..., ge=1),
    db: AsyncSession = Depends(db_session),
    _: None = Depends(require_permission(CONFIG_EDIT)),
) -> PhysicalItemGroupOut:
    group = (
        await db.execute(
            select(PhysicalItemGroup)
            .options(selectinload(PhysicalItemGroup.aliases))
            .where(PhysicalItemGroup.id == group_id)
        )
    ).scalar_one_or_none()
    if group is None:
        raise NotFound(f"库存SKU共享组 {group_id} 不存在")

    new_name = patch.name if patch.name is not None else group.name
    new_members = (
        patch.members if patch.members is not None else [alias.sku for alias in group.aliases]
    )
    await _validate_physical_group_unique(
        db,
        name=new_name,
        members=new_members,
        current_group_id=group.id,
    )

    if patch.name is not None:
        group.name = patch.name
    if patch.enabled is not None:
        group.enabled = patch.enabled
    if "remark" in patch.model_fields_set:
        group.remark = patch.remark
    if patch.members is not None:
        await _replace_physical_members(db, group, patch.members)
    await db.commit()
    group = (
        await db.execute(
            select(PhysicalItemGroup)
            .options(selectinload(PhysicalItemGroup.aliases))
            .where(PhysicalItemGroup.id == group_id)
        )
    ).scalar_one()
    return _physical_group_out(group)


@router.delete("/physical-item-groups/{group_id}", status_code=204)
async def delete_physical_item_group(
    group_id: int = Path(..., ge=1),
    db: AsyncSession = Depends(db_session),
    _: None = Depends(require_permission(CONFIG_EDIT)),
) -> None:
    result = await db.execute(delete(PhysicalItemGroup).where(PhysicalItemGroup.id == group_id))
    if result.rowcount == 0:  # type: ignore[attr-defined]
        raise NotFound(f"库存SKU共享组 {group_id} 不存在")
    await db.commit()
    return None


# ============================================================
# Warehouse
# ============================================================
@router.get("/warehouse", response_model=list[WarehouseOut])
async def list_warehouses(
    db: AsyncSession = Depends(db_session_readonly),
    _: None = Depends(require_permission(CONFIG_VIEW)),
) -> list[WarehouseOut]:
    rows = (await db.execute(_warehouse_list_stmt())).all()
    return [_warehouse_out_from_row(row) for row in rows]


@router.patch("/warehouse/{warehouse_id}/country", response_model=WarehouseOut)
async def patch_warehouse_country(
    patch: WarehouseCountryPatch,
    warehouse_id: str = Path(...),
    db: AsyncSession = Depends(db_session),
    _: None = Depends(require_permission(CONFIG_EDIT)),
) -> WarehouseOut:
    row = (
        await db.execute(select(Warehouse).where(Warehouse.id == warehouse_id))
    ).scalar_one_or_none()
    if row is None:
        raise NotFound(f"仓库 {warehouse_id} 不存在")
    new_country = patch.country.upper() if patch.country else None
    await db.execute(
        update(Warehouse).where(Warehouse.id == warehouse_id).values(country=new_country)
    )
    # 同步更新 inventory_snapshot_latest 中该仓库的 country，避免等下次库存同步
    await db.execute(
        update(InventorySnapshotLatest)
        .where(InventorySnapshotLatest.warehouse_id == warehouse_id)
        .values(country=new_country)
    )
    await db.commit()
    updated_row = (
        await db.execute(_warehouse_list_stmt().where(Warehouse.id == warehouse_id))
    ).one()
    return _warehouse_out_from_row(updated_row)


# ============================================================
# Zipcode Rule
# ============================================================
@router.get("/zipcode-rules", response_model=list[ZipcodeRuleOut])
async def list_zipcode_rules(
    country: str | None = Query(default=None),
    db: AsyncSession = Depends(db_session_readonly),
    _: None = Depends(require_permission(CONFIG_VIEW)),
) -> list[ZipcodeRuleOut]:
    base = select(ZipcodeRule).order_by(ZipcodeRule.country, ZipcodeRule.priority)
    if country:
        base = base.where(ZipcodeRule.country == country.upper())
    rows = (await db.execute(base)).scalars().all()
    return [ZipcodeRuleOut.model_validate(r) for r in rows]


@router.post("/zipcode-rules", response_model=ZipcodeRuleOut, status_code=201)
async def create_zipcode_rule(
    body: ZipcodeRuleIn,
    db: AsyncSession = Depends(db_session),
    _: None = Depends(require_permission(CONFIG_EDIT)),
) -> ZipcodeRuleOut:
    # 校验仓库存在
    wh = (
        await db.execute(select(Warehouse).where(Warehouse.id == body.warehouse_id))
    ).scalar_one_or_none()
    if wh is None:
        raise NotFound(f"仓库 {body.warehouse_id} 不存在")
    if wh.country != body.country.upper():
        raise ValidationFailed(f"仓库 {body.warehouse_id} 与规则国家 {body.country.upper()} 不匹配")

    rule = ZipcodeRule(
        country=body.country.upper(),
        prefix_length=body.prefix_length,
        value_type=body.value_type,
        operator=body.operator,
        compare_value=body.compare_value,
        warehouse_id=body.warehouse_id,
        priority=body.priority,
    )
    db.add(rule)
    await db.flush()
    return ZipcodeRuleOut.model_validate(rule)


@router.patch("/zipcode-rules/{rule_id}", response_model=ZipcodeRuleOut)
async def patch_zipcode_rule(
    body: ZipcodeRuleIn,
    rule_id: int = Path(..., ge=1),
    db: AsyncSession = Depends(db_session),
    _: None = Depends(require_permission(CONFIG_EDIT)),
) -> ZipcodeRuleOut:
    rule = (
        await db.execute(select(ZipcodeRule).where(ZipcodeRule.id == rule_id))
    ).scalar_one_or_none()
    if rule is None:
        raise NotFound(f"规则 {rule_id} 不存在")
    wh = (
        await db.execute(select(Warehouse).where(Warehouse.id == body.warehouse_id))
    ).scalar_one_or_none()
    if wh is None:
        raise NotFound(f"仓库 {body.warehouse_id} 不存在")
    if wh.country != body.country.upper():
        raise ValidationFailed(f"仓库 {body.warehouse_id} 与规则国家 {body.country.upper()} 不匹配")
    await db.execute(
        update(ZipcodeRule)
        .where(ZipcodeRule.id == rule_id)
        .values(
            country=body.country.upper(),
            prefix_length=body.prefix_length,
            value_type=body.value_type,
            operator=body.operator,
            compare_value=body.compare_value,
            warehouse_id=body.warehouse_id,
            priority=body.priority,
        )
    )
    await db.refresh(rule)
    return ZipcodeRuleOut.model_validate(rule)


@router.delete("/zipcode-rules/{rule_id}", status_code=204)
async def delete_zipcode_rule(
    rule_id: int = Path(..., ge=1),
    db: AsyncSession = Depends(db_session),
    _: None = Depends(require_permission(CONFIG_EDIT)),
) -> None:
    res = await db.execute(delete(ZipcodeRule).where(ZipcodeRule.id == rule_id))
    if res.rowcount == 0:  # type: ignore[attr-defined]
        raise NotFound(f"规则 {rule_id} 不存在")
    return None


# ============================================================
# Shop
# ============================================================
@router.get("/shops", response_model=list[ShopOut])
async def list_shops(
    db: AsyncSession = Depends(db_session_readonly),
    _: None = Depends(require_permission(DATA_BASE_VIEW)),
) -> list[ShopOut]:
    rows = (await db.execute(select(Shop).order_by(Shop.id))).scalars().all()
    return [ShopOut.model_validate(r) for r in rows]


@router.post("/shops/refresh")
async def refresh_shops(
    db: AsyncSession = Depends(db_session),
    _: None = Depends(require_permission(SYNC_OPERATE)),
) -> dict[str, Any]:
    """触发 sync_shop 任务从赛狐刷新店铺列表。"""
    task_id, existing = await enqueue_task(db, job_name="sync_shop", trigger_source="manual")
    return {"task_id": task_id, "existing": existing}


@router.patch("/shops/{shop_id}", response_model=ShopOut)
async def patch_shop(
    patch: ShopPatch,
    shop_id: str = Path(...),
    db: AsyncSession = Depends(db_session),
    _: None = Depends(require_permission(DATA_BASE_EDIT)),
) -> ShopOut:
    row = (await db.execute(select(Shop).where(Shop.id == shop_id))).scalar_one_or_none()
    if row is None:
        raise NotFound(f"店铺 {shop_id} 不存在")
    if row.status != "0" and patch.sync_enabled:
        raise ConflictError("授权失效的店铺无法启用同步")
    await db.execute(update(Shop).where(Shop.id == shop_id).values(sync_enabled=patch.sync_enabled))
    await db.refresh(row)
    return ShopOut.model_validate(row)
